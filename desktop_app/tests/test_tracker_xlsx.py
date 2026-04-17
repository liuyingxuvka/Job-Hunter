from __future__ import annotations

import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

DESKTOP_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = DESKTOP_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from jobflow_desktop_app.search.output.tracker_xlsx import (  # noqa: E402
    TRACKER_SHEET_NAME,
    write_tracker_xlsx,
)


class TrackerXlsxTests(unittest.TestCase):
    def _config(self, *, adjacent: bool = False) -> dict:
        return {
            "candidate": {
                "scopeProfile": "adjacent_mbse" if adjacent else "hydrogen_mainline",
            }
        }

    def _job(self, *, url: str = "https://acme.example.com/careers/jobs/12345", date_found: str = "2026-04-14T12:00:00Z") -> dict:
        return {
            "title": "Fuel Cell Reliability Engineer",
            "company": "Acme Hydrogen",
            "location": "Berlin, Germany",
            "url": url,
            "dateFound": date_found,
            "summary": "Hydrogen durability diagnostics role.",
            "companyTags": ["hydrogen", "durability"],
            "listTags": ["推荐"],
            "sourceType": "company",
            "jd": {
                "applyUrl": f"{url}/apply",
                "finalUrl": url,
                "status": 200,
                "ok": True,
                "rawText": "Responsibilities Qualifications Apply now",
            },
            "analysis": {
                "recommend": True,
                "overallScore": 74,
                "matchScore": 74,
                "fitLevelCn": "匹配",
                "fitTrack": "hydrogen_core",
                "jobCluster": "Core-Domain",
                "primaryEvidenceCn": "氢能耐久性与诊断关键词",
                "summaryCn": "氢能耐久性岗位",
                "recommendReasonCn": "与候选人方向相符",
                "adjacentDirectionCn": "",
                "industryClusterCn": "",
            },
        }

    def test_write_persists_current_manual_fields_without_compat_headers(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "jobs_recommended.xlsx"
            job = self._job()
            composite_key = "acme hydrogen|fuel cell reliability engineer|berlin germany"
            manual = {
                job["url"]: {
                    "interest": "感兴趣",
                    "appliedCn": "已投递",
                    "appliedDate": "2026-04-15",
                    "responseStatus": "已回复",
                    "notesCn": "重点关注",
                },
                composite_key: {
                    "interest": "感兴趣",
                    "appliedCn": "已投递",
                    "appliedDate": "2026-04-15",
                    "responseStatus": "已回复",
                    "notesCn": "重点关注",
                },
            }
            write_tracker_xlsx(
                xlsx_path=path,
                jobs=[job],
                manual_by_url=manual,
                config=self._config(),
            )

            workbook = load_workbook(path)
            sheet = workbook[TRACKER_SHEET_NAME]
            headers = [cell.value for cell in sheet[1]]
            self.assertNotIn("兼容_Applied", headers)
            self.assertNotIn("兼容_Notes", headers)
            header_index = {header: idx + 1 for idx, header in enumerate(headers)}
            self.assertEqual(sheet.cell(row=2, column=header_index["感兴趣"]).value, "感兴趣")
            self.assertEqual(sheet.cell(row=2, column=header_index["投递状态"]).value, "已投递")
            self.assertEqual(sheet.cell(row=2, column=header_index["投递日期"]).value, "2026-04-15")
            self.assertEqual(sheet.cell(row=2, column=header_index["跟进状态"]).value, "已回复")
            self.assertEqual(sheet.cell(row=2, column=header_index["备注"]).value, "重点关注")

    def test_write_marks_hidden_columns_and_hyperlinks(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "jobs_recommended.xlsx"
            job = self._job()
            write_tracker_xlsx(
                xlsx_path=path,
                jobs=[job],
                manual_by_url={},
                config=self._config(adjacent=True),
            )

            workbook = load_workbook(path)
            sheet = workbook[TRACKER_SHEET_NAME]
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("副线方向", headers)
            self.assertIn("规范链接", headers)
            canonical_column = headers.index("规范链接") + 1
            url_column = headers.index("职位链接") + 1
            recommend_column = headers.index("推荐") + 1
            recommend_letter = sheet.cell(row=1, column=recommend_column).column_letter
            self.assertTrue(sheet.column_dimensions[recommend_letter].hidden)
            self.assertEqual(
                sheet.cell(row=2, column=url_column).hyperlink.target,
                job["jd"]["applyUrl"],
            )
            self.assertEqual(
                sheet.cell(row=2, column=canonical_column).hyperlink.target,
                job["url"],
            )


if __name__ == "__main__":
    unittest.main()

