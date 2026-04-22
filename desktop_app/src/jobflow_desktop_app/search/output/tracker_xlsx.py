from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..analysis.scoring_contract import overall_score
from .final_output import (
    TRACK_CLUSTER_LABEL,
    TRACK_CN_LABEL,
    build_job_composite_key,
    build_job_dedupe_key,
    canonical_job_url,
    choose_output_job_url,
    infer_region_tag,
    infer_source_quality,
    normalize_job_url,
)
from .manual_tracking import MANUAL_TRACKING_KEYS


TRACKER_SHEET_NAME = "Jobs"

TRACKER_COLUMNS_BASE: list[dict[str, Any]] = [
    {"header": "职位链接", "key": "url", "width": 12},
    {"header": "职位名称", "key": "title", "width": 38},
    {"header": "公司名称", "key": "company", "width": 24},
    {"header": "工作地点", "key": "location", "width": 20},
    {"header": "入表日期", "key": "dateFound", "width": 12},
    {"header": "岗位概述（中文）", "key": "summaryCn", "width": 54},
    {"header": "匹配说明（中文）", "key": "primaryEvidenceCn", "width": 34},
    {"header": "匹配程度", "key": "fitLevelCn", "width": 12},
    {"header": "匹配分", "key": "matchScore", "width": 10},
    {"header": "清单标签", "key": "listTags", "width": 14},
    {"header": "感兴趣", "key": "interest", "width": 12},
    {"header": "投递状态", "key": "appliedCn", "width": 12},
    {"header": "投递日期", "key": "appliedDate", "width": 12},
    {"header": "跟进状态", "key": "responseStatus", "width": 12},
    {"header": "备注", "key": "notesCn", "width": 26},
    {"header": "推荐", "key": "recommend", "width": 10, "hidden": True},
    {"header": "推荐理由（中文）", "key": "recommendReasonCn", "width": 36, "hidden": True},
    {"header": "发布日期", "key": "datePosted", "width": 12, "hidden": True},
    {"header": "匹配轨道", "key": "fitTrack", "width": 20, "hidden": True},
    {"header": "岗位簇", "key": "jobCluster", "width": 20, "hidden": True},
    {"header": "可迁移匹配分", "key": "transferableScore", "width": 14, "hidden": True},
    {"header": "公司标签", "key": "companyTags", "width": 28, "hidden": True},
    {"header": "来源", "key": "source", "width": 24, "hidden": True},
    {"header": "来源质量", "key": "sourceQuality", "width": 16, "hidden": True},
    {"header": "地区标签", "key": "regionTag", "width": 12, "hidden": True},
    {"header": "规范链接", "key": "canonicalUrl", "width": 60, "hidden": True},
    {"header": "岗位原文概述", "key": "summary", "width": 50, "hidden": True},
    {"header": "是否岗位页", "key": "isJobPosting", "width": 12, "hidden": True},
    {"header": "岗位页判断依据", "key": "jobPostingEvidenceCn", "width": 40, "hidden": True},
    {"header": "差距说明（中文）", "key": "gapsCn", "width": 60, "hidden": True},
    {"header": "提问建议（中文）", "key": "questionsCn", "width": 60, "hidden": True},
    {"header": "下一步建议（中文）", "key": "nextActionCn", "width": 30, "hidden": True},
    {"header": "不感兴趣", "key": "notInterested", "width": 12, "hidden": True},
    {"header": "Scope Profile", "key": "scopeProfile", "width": 18, "hidden": True},
]

@dataclass(frozen=True)
class TrackerWorkbookWriteResult:
    path: str
    locked: bool = False


def excel_columns(config: Mapping[str, Any] | None) -> list[dict[str, Any]]:
    del config
    return [dict(column) for column in TRACKER_COLUMNS_BASE]


def to_human_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    return text[:10] if len(text) >= 10 else text


def to_applied_cn_display_value(manual: Mapping[str, Any] | None = None) -> str:
    manual = manual or {}
    return str(manual.get("appliedCn") or "").strip()


def to_response_status_display_value(manual: Mapping[str, Any] | None = None) -> str:
    manual = manual or {}
    return str(manual.get("responseStatus") or "").strip()


def to_notes_cn_display_value(manual: Mapping[str, Any] | None = None) -> str:
    manual = manual or {}
    return str(manual.get("notesCn") or "").strip()


def build_sheet_summary_cn(job: Mapping[str, Any]) -> str:
    analysis = job.get("analysis")
    translation = job.get("translation")
    if not isinstance(analysis, Mapping):
        analysis = {}
    if not isinstance(translation, Mapping):
        translation = {}
    summary_cn = str(analysis.get("summaryCn") or translation.get("summaryCn") or "").strip()
    if summary_cn:
        return summary_cn
    evidence = str(analysis.get("primaryEvidenceCn") or "").strip()
    recommend_reason = str(analysis.get("recommendReasonCn") or "").strip()
    if evidence and recommend_reason:
        return f"{evidence} {recommend_reason}"
    if evidence:
        return evidence
    if recommend_reason:
        return recommend_reason
    return str(job.get("summary") or "").strip()


def parse_boolean_value(value: object) -> bool | None:
    text = str(value or "").strip()
    if not text:
        return None
    lowered = text.lower()
    if text == "是" or lowered in {"true", "yes"}:
        return True
    if text == "否" or lowered in {"false", "no"}:
        return False
    return None


def parse_recommend_value(value: object) -> bool | None:
    return parse_boolean_value(value)


def row_to_job(row: Mapping[str, Any], config: Mapping[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    url = normalize_job_url(row.get("url") or row.get("canonicalUrl") or "")
    if not url:
        return None
    del config
    fit_track = str(row.get("fitTrack") or "").strip() or "direct_fit"

    def _number(text: object) -> int | None:
        raw = str(text or "").strip()
        if not raw:
            return None
        try:
            return int(float(raw))
        except (TypeError, ValueError):
            return None

    def _split(text: object) -> list[str]:
        raw = str(text or "").strip()
        if not raw:
            return []
        return [part.strip() for part in raw.split("|") if part.strip()]

    return {
        "url": url,
        "outputUrl": url,
        "canonicalUrl": normalize_job_url(row.get("canonicalUrl") or "") or url,
        "title": str(row.get("title") or ""),
        "company": str(row.get("company") or ""),
        "companyTags": [part.strip() for part in str(row.get("companyTags") or "").split(",") if part.strip()],
        "location": str(row.get("location") or ""),
        "datePosted": str(row.get("datePosted") or ""),
        "dateFound": str(row.get("dateFound") or ""),
        "source": str(row.get("source") or ""),
        "sourceQuality": str(row.get("sourceQuality") or ""),
        "regionTag": str(row.get("regionTag") or ""),
        "sourceType": str(row.get("sourceType") or ""),
        "listTags": _split(row.get("listTags")),
        "summary": str(row.get("summary") or ""),
        "interest": str(row.get("interest") or ""),
        "appliedDate": str(row.get("appliedDate") or ""),
        "appliedCn": str(row.get("appliedCn") or ""),
        "responseStatus": str(row.get("responseStatus") or ""),
        "notInterested": str(row.get("notInterested") or ""),
        "notesCn": str(row.get("notesCn") or ""),
        "analysis": {
            "summaryCn": str(row.get("summaryCn") or ""),
            "overallScore": _number(row.get("matchScore")),
            "matchScore": _number(row.get("matchScore")),
            "fitLevelCn": str(row.get("fitLevelCn") or ""),
            "fitTrack": fit_track,
            "jobCluster": str(row.get("jobCluster") or TRACK_CLUSTER_LABEL.get(fit_track, TRACK_CLUSTER_LABEL["direct_fit"])),
            "industryTrackCn": TRACK_CN_LABEL.get(fit_track, TRACK_CN_LABEL["direct_fit"]),
            "transferableScore": _number(row.get("transferableScore")),
            "primaryEvidenceCn": str(row.get("primaryEvidenceCn") or ""),
            "adjacentDirectionCn": str(row.get("adjacentDirectionCn") or ""),
            "industryClusterCn": str(row.get("industryClusterCn") or ""),
            "scopeProfile": str(row.get("scopeProfile") or ""),
            "isJobPosting": parse_boolean_value(row.get("isJobPosting")),
            "jobPostingEvidenceCn": str(row.get("jobPostingEvidenceCn") or ""),
            "recommend": parse_recommend_value(row.get("recommend")),
            "recommendReasonCn": str(row.get("recommendReasonCn") or ""),
            "reasonsCn": _split(row.get("reasonsCn")),
            "gapsCn": _split(row.get("gapsCn")),
            "questionsCn": _split(row.get("questionsCn")),
            "nextActionCn": str(row.get("nextActionCn") or ""),
        },
    }


def _apply_list_validation(sheet: Any, column_index: int, allowed_values: list[str]) -> None:
    if column_index <= 0:
        return
    letter = get_column_letter(column_index)
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(allowed_values)}"',
        allow_blank=True,
    )
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}1048576")


def _write_link_cell(cell: Any, target_url: str, text: str) -> None:
    if not target_url:
        cell.value = ""
        return
    cell.value = text
    cell.hyperlink = target_url
    cell.font = Font(color="FF0563C1", underline="single")


def _manual_for_job(job: Mapping[str, Any], manual_by_url: Mapping[str, Mapping[str, Any]]) -> dict[str, str]:
    manual_key = normalize_job_url(job.get("url") or "") or build_job_dedupe_key(job) or build_job_composite_key(job)
    manual_canonical_key = canonical_job_url(job)
    manual_composite_key = build_job_composite_key(job)
    sources = [
        manual_by_url.get(manual_key, {}),
        manual_by_url.get(manual_canonical_key, {}),
        manual_by_url.get(manual_composite_key, {}),
    ]
    merged: dict[str, str] = {}
    for source in sources:
        if not isinstance(source, Mapping):
            continue
        for key in MANUAL_TRACKING_KEYS:
            value = str(source.get(key) or merged.get(key) or "").strip()
            if value:
                merged[key] = value
    return merged


def _joined_text(value: object, sep: str) -> str:
    if isinstance(value, list):
        return sep.join(str(item).strip() for item in value if str(item).strip())
    return str(value or "").strip()


def write_tracker_xlsx(
    *,
    xlsx_path: str | Path,
    jobs: list[Mapping[str, Any]],
    manual_by_url: Mapping[str, Mapping[str, Any]],
    config: Mapping[str, Any] | None,
) -> TrackerWorkbookWriteResult:
    path = Path(xlsx_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = TRACKER_SHEET_NAME
    sheet.freeze_panes = "A2"

    columns = excel_columns(config)
    headers = [column["header"] for column in columns]
    sheet.append(headers)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    scope_profile = ""
    if isinstance(config, Mapping):
        candidate = config.get("candidate")
        if isinstance(candidate, Mapping):
            raw_scope_profiles = candidate.get("scopeProfiles")
            if isinstance(raw_scope_profiles, list):
                scope_profile = " | ".join(
                    str(item or "").strip()
                    for item in raw_scope_profiles
                    if str(item or "").strip()
                )
            if not scope_profile:
                scope_profile = str(candidate.get("scopeProfile") or "").strip()

    for job in jobs:
        analysis = job.get("analysis")
        if not isinstance(analysis, Mapping):
            analysis = {}
        analysis_overall_score = overall_score(analysis)
        manual = _manual_for_job(job, manual_by_url)
        source_quality = str(job.get("sourceQuality") or infer_source_quality(job, config))
        region_tag = str(job.get("regionTag") or infer_region_tag(job))
        canonical_url = canonical_job_url(job) or normalize_job_url(job.get("url") or "")
        final_job_url = choose_output_job_url(job, config) or canonical_url or normalize_job_url(job.get("url") or "")
        tracker_date = to_human_date(job.get("dateFound") or "")

        row_values: dict[str, Any] = {
            "title": str(job.get("title") or ""),
            "company": str(job.get("company") or ""),
            "companyTags": _joined_text(job.get("companyTags"), ", "),
            "jobCluster": str(analysis.get("jobCluster") or ""),
            "fitTrack": str(analysis.get("fitTrack") or ""),
            "transferableScore": analysis.get("transferableScore") or "",
            "primaryEvidenceCn": str(analysis.get("primaryEvidenceCn") or ""),
            "sourceQuality": source_quality,
            "regionTag": region_tag,
            "location": str(job.get("location") or ""),
            "datePosted": to_human_date(job.get("datePosted") or ""),
            "dateFound": tracker_date,
            "adjacentDirectionCn": str(analysis.get("adjacentDirectionCn") or ""),
            "industryClusterCn": str(analysis.get("industryClusterCn") or ""),
            "source": str(job.get("source") or ""),
            "listTags": _joined_text(job.get("listTags"), " | "),
            "summary": str(job.get("summary") or ""),
            "summaryCn": build_sheet_summary_cn(job),
            "matchScore": analysis_overall_score or "",
            "fitLevelCn": str(analysis.get("fitLevelCn") or ""),
            "isJobPosting": "是"
            if analysis.get("isJobPosting") is True
            else "否"
            if analysis.get("isJobPosting") is False
            else "",
            "jobPostingEvidenceCn": str(analysis.get("jobPostingEvidenceCn") or ""),
            "recommend": "是"
            if analysis.get("recommend") is True
            else "否"
            if analysis.get("recommend") is False
            else "",
            "recommendReasonCn": str(analysis.get("recommendReasonCn") or ""),
            "reasonsCn": _joined_text(analysis.get("reasonsCn"), " | "),
            "gapsCn": _joined_text(analysis.get("gapsCn"), " | "),
            "questionsCn": _joined_text(analysis.get("questionsCn"), " | "),
            "nextActionCn": str(analysis.get("nextActionCn") or ""),
            "interest": str(manual.get("interest") or ""),
            "appliedDate": str(manual.get("appliedDate") or ""),
            "appliedCn": to_applied_cn_display_value(manual),
            "responseStatus": to_response_status_display_value(manual),
            "notInterested": str(manual.get("notInterested") or ""),
            "notesCn": to_notes_cn_display_value(manual),
            "scopeProfile": scope_profile,
        }

        row = []
        for column in columns:
            key = column["key"]
            if key == "url":
                row.append(final_job_url)
            elif key == "canonicalUrl":
                row.append(canonical_url)
            else:
                row.append(row_values.get(key, ""))
        sheet.append(row)
        row_index = sheet.max_row
        for column_index, column in enumerate(columns, start=1):
            key = column["key"]
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "url":
                _write_link_cell(cell, final_job_url, "Open Job")
            elif key == "canonicalUrl" and canonical_url:
                _write_link_cell(cell, canonical_url, canonical_url)

    for column_index, column in enumerate(columns, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = float(column.get("width", 12))
        sheet.column_dimensions[letter].hidden = bool(column.get("hidden"))
        sheet.cell(row=1, column=column_index).font = Font(bold=True)
        sheet.cell(row=1, column=column_index).alignment = Alignment(vertical="top", wrap_text=True)

    validations = {
        "interest": ["感兴趣", "一般", "不感兴趣"],
        "appliedCn": ["未投递", "已投递", "面试中", "已拒", "Offer"],
        "responseStatus": ["未回复", "已回复", "面试中", "已拒", "Offer", "已失效"],
        "notInterested": ["否", "是"],
    }
    key_to_index = {column["key"]: index + 1 for index, column in enumerate(columns)}
    for key, options in validations.items():
        _apply_list_validation(sheet, key_to_index.get(key, 0), options)

    try:
        workbook.save(path)
        return TrackerWorkbookWriteResult(path=str(path), locked=False)
    except PermissionError:
        alt_path = path.with_name(path.stem + ".new.xlsx")
        workbook.save(alt_path)
        return TrackerWorkbookWriteResult(path=str(alt_path), locked=True)
